from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='/app/admin/Menu.xlsx')

sheet = wb['Лист1']


def update_from_excel(sheet: Workbook) -> list[dict]:
    """
    Parsing of the Menu.xlsx file to python list of dictionaries
    that fully represent the menu, submenu and dish structure.
    Necessary for further convergence to json and comparison with the pydantic schema models.
    """

    menus: list[dict] = []
    submenus: list[dict] = []
    for row in sheet.iter_rows(min_col=0, max_col=7, values_only=True):

        if row[0]:
            menu_id, title, description, submenus = row[0], row[1], row[2], []
            menu_dict = {
                'id': menu_id,
                'title': title,
                'description': description,
                'submenus': submenus
            }
            menus.append(menu_dict)
        elif row[0] is None and row[1]:
            sub_id, sub_title, sub_description, sub_menu_id = row[1], row[2], row[3], menu_dict['id']
            sub_menu_dict = {
                'id': sub_id,
                'title': sub_title,
                'description': sub_description,
                'menu_id': sub_menu_id,
                'dishes': []
            }
            submenus.append(sub_menu_dict)
        elif row[0] is None and row[1] is None and row[2]:
            d_id, d_title, d_description, d_price = row[2], row[3], row[4], row[5]
            d_discount = row[6]
            if d_discount is None:
                d_discount = 0
            dish_dict = {
                'id': d_id,
                'title': d_title,
                'description': d_description,
                'price': d_price,
                'discount': d_discount,
                'submenu_id': sub_menu_dict['id']
            }
            submenus[-1]['dishes'].append(dish_dict)

    for i in range(len(menus)):
        for j in range(len(submenus)):
            if menus[i]['id'] == submenus[j]['menu_id']:
                menus[i]['submenus'].append(submenus[j])
    return menus
