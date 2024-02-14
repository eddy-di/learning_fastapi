from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelSheetParser:
    """
    Excel parser class, parses from Excel sheet returns dictionaries
    of all menu, submenu and dish instances with their IDs as their keys.
    """

    menus: dict = {}
    submenus: dict = {}
    dishes: dict = {}

    file: Workbook = None
    sheet: Worksheet = None

    def __init__(self, filename: str, sheet_name: str) -> None:
        self.file: Workbook = load_workbook(filename=filename)
        self.sheet: Worksheet = self.file[sheet_name]
        self.menus: dict = {}
        self.submenus: dict = {}
        self.dishes: dict = {}

    def parse(self) -> None:
        """
        Parsing of the Menu.xlsx file to python dict of dictionaries
        that fully represent the menu, submenu and dish structure.
        Necessary for further comparison with the pydantic schema models.
        """

        menus: dict = {}
        submenus: dict = {}
        dishes: dict = {}
        menu_global_id = None
        submenu_global_id = None
        for row in self.sheet.iter_rows(min_col=0, max_col=8, values_only=True):

            if row[0]:
                menu_global_id, title, description = row[0], row[1], row[2]
                menu_dict = {
                    'id': menu_global_id,
                    'title': title,
                    'description': description,
                }
                menus[menu_global_id] = menu_dict

            elif row[0] is None and row[1]:
                submenu_global_id, sub_title, sub_description = row[1], row[2], row[3]
                sub_menu_dict = {
                    'id': submenu_global_id,
                    'title': sub_title,
                    'description': sub_description,
                    'menu_id': menu_global_id,
                }
                submenus[submenu_global_id] = sub_menu_dict

            elif row[0] is None and row[1] is None and row[2]:
                d_id, d_title, d_description, d_price = row[2], row[3], row[4], row[5]
                d_discount = row[6]
                if d_discount is None:
                    d_discount = 0
                dish_dict = {
                    'id': d_id,
                    'title': d_title,
                    'description': d_description,
                    'price': d_price,
                    'discount': d_discount,
                    'submenu_id': submenu_global_id,
                    'menu_id': menu_global_id
                }
                dishes[d_id] = dish_dict

        self.menus = menus
        self.submenus = submenus
        self.dishes = dishes


class JsonParser:
    """
    Json parser class, parses from the result of `menu_preview` endpoint and returns dictionaries
    of all menu, submenu and dish instances with their IDs as their keys.
    """

    menus: dict = {}
    submenus: dict = {}
    dishes: dict = {}
    data: dict = {}

    def __init__(self, json_data: dict) -> None:
        self.data = json_data
        self.menus: dict = {}
        self.submenus: dict = {}
        self.dishes: dict = {}

    def parse(self) -> None:
        """
        Parsing of the result of `get_menus_preview` endpoint to python dict of dictionaries
        that fully represent the menu, submenu and dish structure.
        Necessary for further comparison with the excel parser to compare and do further tasks.
        """

        for menu in self.data:
            if menu['id'] not in self.menus:
                menu_data = {
                    'id': menu['id'],
                    'title': menu['title'],
                    'description': menu['description']
                }
                self.menus[menu['id']] = menu_data

            for submenu in menu['submenus']:
                if submenu['id'] not in self.submenus:
                    subemnu_data = {
                        'id': submenu['id'],
                        'title': submenu['title'],
                        'description': submenu['description'],
                        'menu_id': menu['id']
                    }
                    self.submenus[submenu['id']] = subemnu_data

                for dish in submenu['dishes']:
                    if dish['id'] not in self.dishes:
                        dish_data = {
                            'id': dish['id'],
                            'title': dish['title'],
                            'description': dish['description'],
                            'price': dish['price'],
                            'discount': dish['discount'],
                            'submenu_id': submenu['id'],
                            'menu_id': menu['id']
                        }
                        self.dishes[dish['id']] = dish_data
